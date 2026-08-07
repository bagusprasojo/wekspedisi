import base64
import os
from decimal import Decimal
from html import escape
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

_DLL_DIR_HANDLES = []

def _tenant_logo_path(tenant):
    logo = getattr(tenant, 'logo', None)
    if not logo:
        return ''
    try:
        path = logo.path
    except Exception:
        return ''
    return path if os.path.exists(path) else ''

def _tenant_logo_data_uri(tenant):
    path = _tenant_logo_path(tenant)
    if not path:
        return ''
    mime = 'image/png' if path.lower().endswith('.png') else 'image/jpeg'
    with open(path, 'rb') as image_file:
        data = base64.b64encode(image_file.read()).decode('ascii')
    return f'data:{mime};base64,{data}'

def _text(value):
    return '' if value is None else str(value)

def _money(value):
    if value in (None, ''):
        return ''
    try:
        amount = Decimal(value)
    except Exception:
        return _text(value)
    formatted = f'{amount:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
    return formatted[:-3] if formatted.endswith(',00') else formatted

def _date_id(value):
    return value.strftime('%d/%m/%Y') if hasattr(value, 'strftime') else _text(value)

def _print_datetime_id():
    return timezone.localtime().strftime('%d/%m/%Y %H:%M')

def _tenant_config_value(tenant, kode, default=''):
    if not tenant:
        return default
    try:
        from master.services import get_config_value

        return get_config_value(tenant, kode, required=False) or default
    except Exception:
        return default

def _report_admin_name(tenant):
    return _tenant_config_value(tenant, 'INVOICE_ADMIN_NAME', 'LIA WAHYUNINGSIH')

def _xlsx_filename(filename):
    base, _ = os.path.splitext(filename)
    return f'{base}.xlsx'

def _excel_color(value):
    if not value:
        return 'D9EAF7'
    if isinstance(value, str):
        return value.strip('#').upper()
    return ''.join(f'{int(part * 255):02X}' for part in value)

def _excel_add_logo(ws, tenant):
    path = _tenant_logo_path(tenant)
    if not path:
        return
    try:
        from openpyxl.drawing.image import Image

        image = Image(path)
        image.width = 58
        image.height = 58
        ws.add_image(image, 'A1')
    except Exception:
        return

def _write_xlsx_response(filename, workbook):
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{_xlsx_filename(filename)}"'
    return response

def excel_response(filename, title, headers, rows, tenant=None, number_columns=None):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    number_columns = set(number_columns or [])
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    col_count = len(headers)
    _excel_add_logo(ws, tenant)
    if tenant:
        company_end_col = max(col_count - 1, 2) if col_count > 2 else 2
        date_col = max(col_count, 3) if col_count > 2 else 1
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=company_end_col)
        ws.cell(1, 2, _text(tenant.name)).font = Font(bold=True, size=18)
        ws.cell(1 if col_count > 2 else 4, date_col, f'Date: {_print_datetime_id()}').alignment = Alignment(horizontal='right')
        ws.cell(2, 2, f'Office : {_text(tenant.address)}, Kabupaten {_text(tenant.city)}')
        ws.cell(3, 2, f'{_text(tenant.postal_code)}, {_text(tenant.province)}')
        title_row = 5
    else:
        title_row = 1
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=max(col_count, 1))
    ws.cell(title_row, 1, title).font = Font(bold=True, size=14)
    ws.cell(title_row, 1).alignment = Alignment(horizontal='center')
    thin = Side(style='thin', color='94A3B8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill('solid', fgColor='E5E7EB')
    start_row = title_row + 2
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row_no, row in enumerate(rows, start=start_row + 1):
        for col, value in enumerate(row, start=1):
            is_number = isinstance(value, (int, float, Decimal))
            is_number_column = (col - 1) in number_columns
            cell = ws.cell(row_no, col, value if is_number else _text(value))
            cell.border = border
            if is_number or is_number_column:
                cell.number_format = '#,##0.##'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(wrap_text=True)
    for col in range(1, len(headers) + 1):
        width = max(len(_text(ws.cell(row, col).value)) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 40)
    return _write_xlsx_response(filename, wb)

def legacy_report_excel_response(filename, title, tenant, period, headers, rows, totals, extra_lines=None, header_color='#b6d2e9'):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    extra_lines = extra_lines or []
    admin_name = _report_admin_name(tenant)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    col_count = len(headers)
    _excel_add_logo(ws, tenant)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=max(col_count - 1, 2))
    ws.cell(1, 2, _text(tenant.name)).font = Font(bold=True, size=18)
    ws.cell(1, col_count, f'Date: {_print_datetime_id()}').alignment = Alignment(horizontal='right')
    ws.cell(2, 2, f'Office : {_text(tenant.address)}, Kabupaten {_text(tenant.city)}')
    ws.cell(3, 2, f'{_text(tenant.postal_code)}, {_text(tenant.province)}')
    title_row = 5
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=col_count)
    ws.cell(title_row, 1, title).font = Font(bold=True, size=16)
    ws.cell(title_row, 1).alignment = Alignment(horizontal='center')
    row_no = title_row + 1
    for line in extra_lines:
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=col_count)
        ws.cell(row_no, 1, line).alignment = Alignment(horizontal='center')
        row_no += 1
    ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=col_count)
    ws.cell(row_no, 1, f'Periode : {period}').alignment = Alignment(horizontal='center')
    row_no += 2

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill('solid', fgColor=_excel_color(header_color))
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row_no, col, header)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row_no += 1
    for row in rows:
        for col, (value, kind) in enumerate(row, start=1):
            display = value if kind == 'number' else _date_id(value) if kind == 'date' else _text(value)
            cell = ws.cell(row_no, col, display)
            cell.border = border
            cell.alignment = Alignment(horizontal='right' if kind == 'number' else 'center' if kind == 'center' else 'left', wrap_text=True)
            if kind == 'number':
                cell.number_format = '#,##0.##'
        row_no += 1
    col = 1
    for item in totals:
        if len(item) == 5:
            _, _, value, kind, span = item
        else:
            value, kind, span = item
        if span > 1:
            ws.merge_cells(start_row=row_no, start_column=col, end_row=row_no, end_column=col + span - 1)
        cell = ws.cell(row_no, col, value if kind == 'number' else _text(value))
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='right')
        if kind == 'number':
            cell.number_format = '#,##0.##'
        for merged_col in range(col + 1, col + span):
            ws.cell(row_no, merged_col).border = border
        col += span
    row_no += 3
    ws.merge_cells(start_row=row_no, start_column=max(col_count - 3, 1), end_row=row_no, end_column=col_count)
    ws.cell(row_no, max(col_count - 3, 1), f'Kab. Sukoharjo, {_print_datetime_id()}').alignment = Alignment(horizontal='right')
    row_no += 1
    ws.cell(row_no, max(col_count - 5, 1), 'Mengetahui,').alignment = Alignment(horizontal='center')
    ws.cell(row_no, max(col_count - 1, 1), 'Yang membuat').alignment = Alignment(horizontal='center')
    row_no += 4
    ws.cell(row_no, max(col_count - 5, 1), '-').alignment = Alignment(horizontal='center')
    ws.cell(row_no, max(col_count - 1, 1), admin_name).alignment = Alignment(horizontal='center')
    row_no += 1
    ws.cell(row_no, max(col_count - 5, 1), 'Manager Operasional').font = Font(bold=True)
    ws.cell(row_no, max(col_count - 1, 1), 'Admin').font = Font(bold=True)
    for column in range(1, col_count + 1):
        width = max(len(_text(ws.cell(row, column).value)) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 9), 32)
    return _write_xlsx_response(filename, wb)

def _rows_html(headers, rows, totals=None, header_color='#e5e7eb', col_widths=None):
    colgroup = ''
    if col_widths:
        width_sum = sum(col_widths) or 1
        colgroup = '<colgroup>' + ''.join(f'<col style="width:{width / width_sum * 100:.2f}%">' for width in col_widths) + '</colgroup>'
    html = [f'<table class="data">{colgroup}<thead><tr>']
    for header in headers:
        html.append(f'<th style="background:{header_color}">{escape(_text(header))}</th>')
    html.append('</tr></thead><tbody>')
    for row in rows:
        html.append('<tr>')
        for value, kind in row:
            css = 'num' if kind == 'number' else 'center' if kind == 'center' else ''
            display = _text(value) if kind == 'number' and isinstance(value, str) else _money(value) if kind == 'number' else _date_id(value) if kind == 'date' else _text(value)
            html.append(f'<td class="{css}">{escape(display)}</td>')
        html.append('</tr>')
    if totals:
        html.append('<tr class="total">')
        for item in totals:
            if len(item) == 5:
                _, _, value, kind, span = item
            else:
                value, kind, span = item
            css = 'num' if kind == 'number' else 'label'
            display = _text(value) if kind == 'number' and isinstance(value, str) else _money(value) if kind == 'number' else _text(value)
            html.append(f'<td class="{css}" colspan="{span}">{escape(display)}</td>')
        html.append('</tr>')
    html.append('</tbody></table>')
    return ''.join(html)

def _pdf_html_document(title, tenant, body, landscape=False):
    logo = _tenant_logo_data_uri(tenant)
    logo_html = f'<img class="logo" src="{logo}">' if logo else '<div class="logo"></div>'
    page_size = 'A4 landscape' if landscape else 'A4'
    tenant_name = getattr(tenant, 'name', '')
    tenant_address = getattr(tenant, 'address', '')
    tenant_city = getattr(tenant, 'city', '')
    tenant_postal_code = getattr(tenant, 'postal_code', '')
    tenant_province = getattr(tenant, 'province', '')
    return f'''<!doctype html>
<html>
<head>
<meta charset="utf-8">
<style>
@page {{ size: {page_size}; margin: 14mm 10mm; }}
body {{ font-family: Arial, sans-serif; font-size: 9pt; color: #000; }}
.company-header {{ display: grid; grid-template-columns: 58px 1fr auto; gap: 10px; align-items: start; border-bottom: 1px solid #000; padding-bottom: 8px; }}
.logo {{ width: 50px; height: 50px; object-fit: contain; }}
.company {{ font-size: 19pt; font-weight: 700; line-height: 1; }}
.addr {{ margin-top: 4px; }}
.date {{ text-align: right; white-space: nowrap; }}
h1 {{ font-size: 16pt; margin: 22px 0 8px; text-align: center; }}
.subtitle {{ text-align: center; margin: 3px 0; }}
table.data {{ width: 100%; border-collapse: collapse; margin-top: 14px; table-layout: fixed; }}
.data th, .data td {{ border: 1px solid #000; padding: 4px; vertical-align: middle; word-wrap: break-word; overflow-wrap: anywhere; }}
.data th {{ text-align: center; font-weight: 700; }}
.num {{ text-align: right; white-space: nowrap; }}
.center {{ text-align: center; }}
.label {{ text-align: right; font-weight: 700; }}
.total td {{ font-weight: 700; }}
.sign {{ width: 100%; margin-top: 28px; page-break-inside: avoid; }}
.sign td {{ border: 0; text-align: center; height: 18px; }}
.sign .date-line {{ text-align: right; }}
</style>
</head>
<body>
<div class="company-header">
{logo_html}
<div>
<div class="company">{escape(_text(tenant_name))}</div>
<div class="addr">Office : {escape(_text(tenant_address))}, Kabupaten {escape(_text(tenant_city))}</div>
<div>{escape(_text(tenant_postal_code))}, {escape(_text(tenant_province))}</div>
</div>
<div class="date">Date: {escape(_print_datetime_id())}</div>
</div>
<h1>{escape(title)}</h1>
{body}
</body>
</html>'''

def _weasy_response(filename, html, inline=False):
    if os.name == 'nt':
        dll_dir = r'C:\msys64\ucrt64\bin'
        if os.path.isdir(dll_dir) and hasattr(os, 'add_dll_directory'):
            _DLL_DIR_HANDLES.append(os.add_dll_directory(dll_dir))
    from weasyprint import HTML

    response = HttpResponse(HTML(string=html).write_pdf(), content_type='application/pdf')
    disposition = 'inline' if inline else 'attachment'
    response['Content-Disposition'] = f'{disposition}; filename="{filename}"'
    return response

def pdf_response(filename, title, headers, rows, landscape=True, tenant=None, number_columns=None, col_widths=None):
    number_columns = set(number_columns or [])
    typed_rows = [
        [
            (value, 'number' if col in number_columns or isinstance(value, (int, float, Decimal)) else 'text')
            for col, value in enumerate(row)
        ]
        for row in rows
    ]
    html = _pdf_html_document(title, tenant, _rows_html(headers, typed_rows, col_widths=col_widths), landscape=landscape)
    return _weasy_response(filename, html)

def legacy_report_pdf_response(filename, title, tenant, period, columns, rows, totals, extra_lines=None, header_rgb=(0.71, 0.82, 0.91), title_italic=False, landscape=False):
    extra_lines = extra_lines or []
    admin_name = _report_admin_name(tenant)
    subtitles = ''.join(f'<div class="subtitle">{escape(line)}</div>' for line in extra_lines)
    subtitles += f'<div class="subtitle">Periode : {escape(period)}</div>'
    body = subtitles + _rows_html(
        [col['label'].replace('\n', ' ') for col in columns],
        rows,
        totals,
        header_color='#' + _excel_color(header_rgb),
        col_widths=[col['w'] for col in columns],
    )
    body += f'''
<table class="sign">
<tr><td colspan="4" class="date-line">Kab. Sukoharjo, {escape(_print_datetime_id())}</td></tr>
<tr><td></td><td>Mengetahui,</td><td></td><td>Yang membuat</td></tr>
<tr><td colspan="4" style="height:46px"></td></tr>
<tr><td></td><td style="border-bottom:1px solid #000">-</td><td></td><td style="border-bottom:1px solid #000">{escape(admin_name)}</td></tr>
<tr><td></td><td><strong>Manager Operasional</strong></td><td></td><td><strong>Admin</strong></td></tr>
</table>'''
    return _weasy_response(filename, _pdf_html_document(title, tenant, body, landscape=landscape))
