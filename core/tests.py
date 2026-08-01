from io import BytesIO

from django.test import SimpleTestCase, TestCase

from core.exporters import _report_admin_name, _rows_html, excel_response, legacy_report_excel_response, legacy_report_pdf_response, pdf_response
from master.models import TenantConfig
from tenants.models import Tenant

class ExporterTests(SimpleTestCase):
    def test_excel_response_converts_non_number_objects_to_text(self):
        class Value:
            def __str__(self):
                return '5010315 - BIAYA PEMBELIAN SPAREPART'

        response = excel_response('test.xls', 'Test', ['Akun'], [[Value()]])

        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertTrue(response.content.startswith(b'PK'))

    def test_excel_response_sets_width_with_merged_title_cells(self):
        response = excel_response('test.xls', 'Test', ['No', 'Akun', 'Nominal'], [[1, 'Kas', 1000]])

        self.assertTrue(response.content.startswith(b'PK'))

    def test_excel_response_prints_tenant_header(self):
        from openpyxl import load_workbook

        tenant = Tenant(name='PT Tampir Baru Logistik', address='Jl Test', city='Sukoharjo')
        response = excel_response('test.xls', 'Transaksi Kas', ['No', 'Akun'], [[1, 'Kas']], tenant=tenant)
        ws = load_workbook(BytesIO(response.content)).active

        self.assertEqual(ws['B1'].value, 'PT Tampir Baru Logistik')
        self.assertRegex(ws['A4'].value, r'^Date: \d{2}/\d{2}/\d{4} \d{2}:\d{2}$')
        self.assertTrue(ws['B2'].value.startswith('Office : Jl Test'))
        self.assertEqual(ws['A5'].value, 'Transaksi Kas')

    def test_rows_html_accepts_legacy_pdf_total_shape(self):
        html = _rows_html(
            ['No', 'Keluar'],
            [[(1, 'number'), (1000, 'number')]],
            [(0, 50, 'Total', 'text', 1), (50, 50, 1000, 'number', 1)],
        )

        self.assertIn('Total', html)
        self.assertIn('1.000', html)

    def test_pdf_response_marks_configured_string_money_column_as_number(self):
        captured = {}

        def fake_weasy(filename, html, inline=False):
            captured['html'] = html
            return filename

        import core.exporters

        original = core.exporters._weasy_response
        core.exporters._weasy_response = fake_weasy
        try:
            pdf_response('test.pdf', 'Test', ['Nama', 'Nominal BBM'], [['Solar', '1.000']], number_columns=[1])
        finally:
            core.exporters._weasy_response = original

        self.assertIn('<td class="num">1.000</td>', captured['html'])

    def test_pdf_response_uses_column_widths(self):
        captured = {}

        def fake_weasy(filename, html, inline=False):
            captured['html'] = html
            return filename

        import core.exporters

        original = core.exporters._weasy_response
        core.exporters._weasy_response = fake_weasy
        try:
            pdf_response('test.pdf', 'Test', ['No Bukti', 'Debet'], [['BNK-001', '1.000']], col_widths=[2, 1], number_columns=[1])
        finally:
            core.exporters._weasy_response = original

        self.assertIn('<col style="width:66.67%">', captured['html'])
        self.assertIn('<col style="width:33.33%">', captured['html'])

    def test_excel_response_aligns_configured_string_money_column_right(self):
        from openpyxl import load_workbook

        response = excel_response('test.xls', 'Test', ['Nama', 'Nominal BBM'], [['Solar', '1.000']], number_columns=[1])
        ws = load_workbook(BytesIO(response.content)).active

        self.assertEqual(ws['B4'].alignment.horizontal, 'right')

class TenantReportConfigTests(TestCase):
    def test_report_admin_name_uses_tenant_config(self):
        tenant = Tenant.objects.create(name='PT Test')
        TenantConfig.objects.create(tenant=tenant, kode='INVOICE_ADMIN_NAME', nilai='Admin Operasional')

        self.assertEqual(_report_admin_name(tenant), 'Admin Operasional')

    def test_legacy_report_excel_uses_configured_admin_name(self):
        from openpyxl import load_workbook

        tenant = Tenant.objects.create(name='PT Test')
        TenantConfig.objects.create(tenant=tenant, kode='INVOICE_ADMIN_NAME', nilai='Admin Operasional')
        response = legacy_report_excel_response(
            'laporan.xls',
            'Laporan',
            tenant,
            '01/07/2026 s.d. 31/07/2026',
            ['No', 'Nilai'],
            [[(1, 'number'), (1000, 'number')]],
            [('Total', 'text', 1), (1000, 'number', 1)],
        )
        ws = load_workbook(BytesIO(response.content)).active

        self.assertIn('Admin Operasional', [cell.value for row in ws.iter_rows() for cell in row])

    def test_legacy_report_pdf_uses_configured_admin_name(self):
        captured = {}

        def fake_weasy(filename, html, inline=False):
            captured['html'] = html
            return filename

        import core.exporters

        tenant = Tenant.objects.create(name='PT Test')
        TenantConfig.objects.create(tenant=tenant, kode='INVOICE_ADMIN_NAME', nilai='Admin Operasional')
        original = core.exporters._weasy_response
        core.exporters._weasy_response = fake_weasy
        try:
            legacy_report_pdf_response(
                'laporan.pdf',
                'Laporan',
                tenant,
                '01/07/2026 s.d. 31/07/2026',
                [{'label': 'No', 'w': 20}],
                [[(1, 'number')]],
                [],
            )
        finally:
            core.exporters._weasy_response = original

        self.assertIn('Admin Operasional', captured['html'])
